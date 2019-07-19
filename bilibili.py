import urllib.request 
import random         
import time
import json
import xlrd
import xlwt
import openpyxl


def requests_headers():
    #构造请求头池 
    head_connection = ['Keep-Alive','close'] 
    head_accept = ['text/html,application/xhtml+xml,*/*']
    head_accept_language = ['zh-CN,fr-FR;q=0.5', 'en-US,en;q=0.8,zh-Hans-CN;q=0.5,zh-Hans;q=0.3']
    head_user_agent = ['Mozilla/5.0 (Windows NT 6.3; WOW64; Trident/7.0; rv:11.0) like Gecko', 'Mozilla/5.0 (Windows NT 5.1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/28.0.1500.95 Safari/537.36', 'Mozilla/5.0 (Windows NT 6.1; WOW64; Trident/7.0; SLCC2; .NET CLR 2.0.50727; .NET CLR 3.5.30729; .NET CLR 3.0.30729; Media Center PC 6.0; .NET4.0C; rv:11.0) like Gecko)', 'Mozilla/5.0 (Windows; U; Windows NT 5.2) Gecko/2008070208 Firefox/3.0.1', 'Mozilla/5.0 (Windows; U; Windows NT 5.1) Gecko/20070309 Firefox/2.0.0.3', 'Mozilla/5.0 (Windows; U; Windows NT 5.1) Gecko/20070803 Firefox/1.5.0.12', 'Opera/9.27 (Windows NT 5.2; U; zh-cn)', 'Mozilla/5.0 (Macintosh; PPC Mac OS X; U; en) Opera 8.0', 'Opera/8.0 (Macintosh; PPC Mac OS X; U; en)', 'Mozilla/5.0 (Windows; U; Windows NT 5.1; en-US; rv:1.8.1.12) Gecko/20080219 Firefox/2.0.0.12 Navigator/9.0.0.6', 'Mozilla/4.0 (compatible; MSIE 8.0; Windows NT 6.1; Win64; x64; Trident/4.0)', 'Mozilla/4.0 (compatible; MSIE 8.0; Windows NT 6.1; Trident/4.0)', 'Mozilla/5.0 (compatible; MSIE 10.0; Windows NT 6.1; WOW64; Trident/6.0; SLCC2; .NET CLR 2.0.50727; .NET CLR 3.5.30729; .NET CLR 3.0.30729; Media Center PC 6.0; InfoPath.2; .NET4.0C; .NET4.0E)', 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.1 (KHTML, like Gecko) Maxthon/4.0.6.2000 Chrome/26.0.1410.43 Safari/537.1 ', 'Mozilla/5.0 (compatible; MSIE 10.0; Windows NT 6.1; WOW64; Trident/6.0; SLCC2; .NET CLR 2.0.50727; .NET CLR 3.5.30729; .NET CLR 3.0.30729; Media Center PC 6.0; InfoPath.2; .NET4.0C; .NET4.0E; QQBrowser/7.3.9825.400)', 'Mozilla/5.0 (Windows NT 6.1; WOW64; rv:21.0) Gecko/20100101 Firefox/21.0 ', 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.1 (KHTML, like Gecko) Chrome/21.0.1180.92 Safari/537.1 LBBROWSER', 'Mozilla/5.0 (compatible; MSIE 10.0; Windows NT 6.1; WOW64; Trident/6.0; BIDUBrowser 2.x)', 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/536.11 (KHTML, like Gecko) Chrome/20.0.1132.11 TaoBrowser/3.0 Safari/536.11']
    header = {'Connection': head_connection[random.randrange(0, len(head_connection))], 'Accept': head_accept[0], 'Accept-Language': head_accept_language[random.randrange(0, len(head_accept_language))], 'User-Agent': head_user_agent[random.randrange(0, len(head_user_agent))],}
    #获得随机请求头 
    return header


#代理IP池 
proxies = ['125.66.217.114:6675','112.251.161.82:6675', '117.34.253.157:6675','113.94.72.209:6666', '114.105.217.144:6673','125.92.110.80:6675', '112.235.126.55:6675','14.148.99.188:6675', '112.240.161.20:6668','122.82.160.148:6675', '175.30.224.66:6675']

#proxies = ['61.135.217.7:80','202.121.178.244:80','114.115.216.99:80']


#抽取IP池IP，构建opener 
def request_proxie():
    header1 = requests_headers()  # 获得随机请求头 
    proxie_handler = urllib.request.ProxyHandler({'http': random.choice(proxies)})
    opener = urllib.request.build_opener(proxie_handler)
    header = []
    for key, value in header1.items():
        elem = (key, value)
        header.append(elem)
    opener.addheaders = header
    return opener


def getHtmlInfo(av_num):
    opener = request_proxie()
    req_view = opener.open("https://api.bilibili.com/x/web-interface/view?aid=" + str(av_num))
    page_view = req_view.read().decode('utf-8')
    dic_page = json.loads(page_view)  # 将获取内容转成字典形式 
    if dic_page['code'] == 0:
        video_info = dic_page['data']
        if 'vlog' in video_info['title'] or 'VLOG' in video_info['title'] or 'Vlog' in video_info['title']:
            video_data = [video_info['aid'],  #av号 
                        video_info['title'],  #标题 
                        video_info['tname'],  #视频类型 
                        time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(video_info['ctime'])),
                        video_info['owner']['name'],  #作者
                        video_info['stat']['view'],  #阅读量
                        video_info['stat']['danmaku'],  #弹幕数 
                        video_info['stat']['reply'],  #评论数 
                        video_info['stat']['favorite'],  #收藏数 
                        video_info['stat']['coin'],  #投币数 
                        video_info['stat']['like']  #点赞数 
                        ]
            time.sleep(0.4)
            #print(video_data)
            print("爬取av" + str(av_num) + '已完成')  #提示爬取进度 
            req_view.close()
            return video_data
    else:
        time.sleep(0.5)


def getAllInfo(start_Num, end_Num):
    allinfo = []
    for av_num in range(start_Num, end_Num + 1):
        tempinfo = getHtmlInfo(av_num)
        if(tempinfo!=None):
            allinfo.append(tempinfo)
    
    print(len(allinfo))
    #print(allinfo)
    return allinfo

def set_style(self,name,height,bold=False):
    style = xlwt.XFStyle() # 初始化样式
    font = xlwt.Font() # 为样式创建字体
    font.name = name  # 'Times New Roman'
    font.bold = bold
    font.color_index = 4
    font.height = height
    borders= xlwt.Borders()
    borders.left= 6
    borders.right= 6
    borders.top= 6
    borders.bottom= 6
    style.font = font
    style.borders = borders
    return style


def savetoExcell(start_Num, end_Num):
    allinfo = getAllInfo(start_Num, end_Num)

    #wb = openpyxl.Workbook()
    wb = xlwt.Workbook()

    count = 1
    #sheet1 = wb.create_sheet('bilibili视频信息爬取', 0)
    sheet1 = wb.add_sheet(u'sheet1',cell_overwrite_ok = True)

    title_ls = ['av号', '视频标题', '视频类型', '上传时间', '作者（UP主）',' 阅读量','弹幕数', '评论数', '收藏数', '投币数', '点赞数']
    for i in range(len(title_ls)):
        sheet1.write(0, i, title_ls[i])
    if len(allinfo)>0:
        for row in range(len(allinfo)):
            #print('真正写入av' + str(allinfo[0][0] + row))  #提示写入进度
            if(len(allinfo[row])>0):
                for column in range(len(allinfo[row])):
                    sheet1.write(count,column, allinfo[row][column])
                count+=1

    #
    # for i in range(len(title_ls)):  #循环写入标题 
    #    sheet1.cell(1, 1 + i).value = title_ls[i]
    #for row in range(len(allinfo)):  #嵌套循环写入视频信息 
    #    print('真正写入av' + str(allinfo[0][0] + row))  #提示写入进度 
    #    for column in range(len(allinfo[row])):
    #        sheet1.cell(row + 2, column + 1).value = allinfo[row][column]
            
    wb.save('D:/bilibili_data' + '_' + str(start_Num) + '_' + str(end_Num) + '.xls')
    #wb.close() 


if __name__ == '__main__':
    for i in range(51323102, 51333102,10000):
        savetoExcell(i,i+10000)


